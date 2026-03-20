import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

FILE = "test_startups.xlsx"

def create_test_db():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Startups"

    headers = [
        "Company Name", "Sector", "Funding Round", "Amount Raised",
        "Date Funded", "Founder Name", "Founder Email", "Website",
        "Source URL", "Email Status", "Email Sent Date", "Reply Date",
        "Notes", "Company Email",
    ]

    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill(start_color="4B0082", end_color="4B0082", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    col_widths = {1:25,2:15,3:15,4:15,5:15,6:20,7:30,8:30,9:40,10:18,11:18,12:15,13:30,14:30}
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # ── TEST ROWS ─────────────────────────────────────────────────────────────
    # Row 1: has founder email      → sender should use founder email
    # Row 2: has only company email → sender should fall back to company email
    # Row 3: no email at all        → sender should skip
    # Row 4: already sent           → sender should skip
    # Row 5: already replied        → sender should skip

    rows = [
        ["TestCo Alpha",   "D2C",         "Series A", "$5M",  "2025-11-01", "Harsh Test",  "harshborude11@gmail.com", "https://testco.com",   "", "Not sent",      "", "", "Test row 1 — has founder email",      ""],
        ["TestCo Beta",    "Fintech",      "Series B", "$12M", "2025-10-15", "",            "",                        "https://testbeta.com", "", "Not sent",      "", "", "Test row 2 — has only company email", "harshborude11@gmail.com"],
        ["TestCo Gamma",   "Edtech",       "Series A", "$3M",  "2025-09-20", "",            "",                        "",                     "", "Not sent",      "", "", "Test row 3 — no email, should skip",  ""],
        ["TestCo Delta",   "Consumer Tech","Series B", "$8M",  "2025-08-10", "Priya Sharma","priya@testdelta.com",     "https://testdelta.com","", "Sent – no reply","2026-03-01","","Test row 4 — already sent, skip", ""],
        ["TestCo Epsilon", "D2C",          "Series A", "$6M",  "2025-07-05", "Arjun Mehta", "arjun@testepsilon.com",  "https://testeps.com",  "", "Replied",       "2026-02-15","2026-02-18","Test row 5 — replied, skip",""],
    ]

    for row_data in rows:
        ws.append(row_data)

    wb.save(FILE)
    print(f"Created {FILE} with {len(rows)} test rows.")
    print()
    print("Test rows summary:")
    print("  Row 2 — TestCo Alpha  : will send to harshborude11@gmail.com (founder email)")
    print("  Row 3 — TestCo Beta   : will send to harshborude11@gmail.com (company email fallback)")
    print("  Row 4 — TestCo Gamma  : SKIP — no email")
    print("  Row 5 — TestCo Delta  : SKIP — already sent")
    print("  Row 6 — TestCo Epsilon: SKIP — already replied")
    print()
    print("Expected: exactly 2 emails sent to harshborude11@gmail.com")

if __name__ == "__main__":
    create_test_db()