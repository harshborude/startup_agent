import requests
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import time
import re
import openpyxl
import os

FILE = "startups.xlsx"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
}

FUNDING_KEYWORDS = ["series a", "series b", "raises", "raise", "raised", "secures", "secured", "bags", "bagged"]

B2B_KEYWORDS = [
    "enterprise", "b2b", "procurement", "supply chain", "manufacturing",
    "infrastructure", "developer tools", "cybersecurity", "devops",
    "erp", "crm", "hrtech", "hr tech", "deep tech", "agritech",
    "proptech", "spacetech", "climatetech", "defence", "defense",
]

# Roundup article patterns to skip entirely
ROUNDUP_PATTERNS = [
    r"from .+ to .+",
    r"weekly funding",
    r"monthly funding",
    r"funding report",
    r"funding roundup",
    r"this week",
    r"startups that raised",
    r"indian startups",
    r"india startups",
    r"top \d+ startups",
]

def is_roundup(text):
    text_lower = text.lower()
    return any(re.search(p, text_lower) for p in ROUNDUP_PATTERNS)

def is_b2b(text):
    text_lower = text.lower()
    return any(kw in text_lower for kw in B2B_KEYWORDS)

def is_funding_article(text):
    text_lower = text.lower()
    return any(kw in text_lower for kw in FUNDING_KEYWORDS)

def is_valid_company_name(name):
    """Filter out junk company names."""
    if not name or len(name) < 3 or len(name) > 60:
        return False
    # Must start with a letter
    if not name[0].isalpha():
        return False
    # No slashes, no URLs
    if "/" in name or "http" in name.lower():
        return False
    # Not a sentence (too many spaces = probably a headline, not a name)
    if len(name.split()) > 6:
        return False
    # Not a roundup
    if is_roundup(name):
        return False
    return True

def extract_amount(text):
    match = re.search(r'\$[\d,.]+\s*(?:million|mn|billion|bn)?', text, re.IGNORECASE)
    if match:
        return match.group(0).strip()
    match = re.search(r'Rs\.?\s*[\d,.]+\s*(?:crore|lakh|cr)', text, re.IGNORECASE)
    if match:
        return match.group(0).strip()
    match = re.search(r'[\d,.]+\s*(?:crore|cr)\b', text, re.IGNORECASE)
    if match:
        return "Rs. " + match.group(0).strip()
    return "N/A"

def extract_round(text):
    text_lower = text.lower()
    if "series b" in text_lower:
        return "Series B"
    if "series a" in text_lower:
        return "Series A"
    return "Unknown"

def extract_company_from_title(title):
    """
    Extract company name from a funding headline like:
    'Burger Singh raises $5M in Series A'
    'Lenskart secures $100M Series B funding'
    """
    # Skip roundup articles immediately
    if is_roundup(title):
        return ""

    # Pattern: <Company> raises/secures/bags/gets <amount>
    match = re.match(
        r'^(.+?)\s+(?:raises|raise|raised|secures|secured|bags|bagged|gets|closes|closed|lands)\s+',
        title, re.IGNORECASE
    )
    if match:
        name = match.group(1).strip()
        # Clean common prefixes like "Edtech startup Lenskart" -> "Lenskart"
        # Remove descriptors: "Edtech startup X" -> "X"
        name = re.sub(
            r'^(?:edtech|fintech|healthtech|d2c|foodtech|gaming|travel|fashion|'
            r'beauty|wellness|fitness|media|retail|consumer|mobility|insurtech|'
            r'logistics|payments|lending|startup|company|firm|platform|app|'
            r'brand|unicorn|soonicorn)\s+(?:startup\s+|firm\s+|company\s+|platform\s+)?',
            '', name, flags=re.IGNORECASE
        ).strip()
        return name

    return ""


def extract_company_from_slug(slug):
    """
    Extract company name from a URL slug like:
    /news/burger-singh-raises-5-mn-series-a-12345
    """
    # Remove trailing numeric ID
    slug = re.sub(r'-\d+$', '', slug)
    # Remove /news/ prefix
    slug = slug.replace("/news/", "").strip("/")

    # Split on action verb in slug
    parts = re.split(r'-(?:raises|secures|bags|gets|closes|lands)-', slug, flags=re.IGNORECASE)
    if len(parts) >= 2:
        company = parts[0].replace("-", " ").title().strip()
        return company

    return ""

def load_existing_companies():
    if not os.path.exists(FILE):
        return set()
    wb = openpyxl.load_workbook(FILE)
    ws = wb.active
    companies = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            companies.add(str(row[0]).strip().lower())
    return companies

def save_to_excel(startups):
    if not startups:
        print("No new startups to save.")
        return 0

    wb = openpyxl.load_workbook(FILE)
    ws = wb.active
    existing = load_existing_companies()
    added = 0

    for s in startups:
        name = s.get("company", "").strip()
        if not is_valid_company_name(name):
            continue
        if name.lower() in existing:
            continue

        ws.append([
            name,
            s.get("sector", "Consumer Tech"),
            s.get("round", ""),
            s.get("amount", ""),
            s.get("date", datetime.now().strftime("%Y-%m-%d")),
            "",           # Founder Name
            "",           # Founder Email
            s.get("website", ""),
            s.get("source_url", ""),
            "Not sent",   # Email Status
            "",           # Email Sent Date
            "",           # Reply Date
            "",           # Notes
        ])
        existing.add(name.lower())
        added += 1

    wb.save(FILE)
    return added


# ── ENTRACKR ──────────────────────────────────────────────────────────────────

def scrape_entrackr():
    print("Scraping Entrackr...")
    results = []

    try:
        resp = requests.get("https://entrackr.com", headers=HEADERS, timeout=15)
        soup = BeautifulSoup(resp.text, "html.parser")

        seen_urls = set()
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if not href.startswith("/news/"):
                continue
            if href in seen_urls:
                continue
            seen_urls.add(href)

            slug = href.lower()
            # Only process slugs that contain funding keywords
            if not any(kw in slug for kw in ["raises", "secures", "bags", "series-a", "series-b", "funding"]):
                continue

            title = a.get_text(strip=True)
            # Use slug-based extraction if title is missing or too short
            if not title or len(title) < 15:
                company = extract_company_from_slug(href)
            else:
                company = extract_company_from_title(title)
                if not company:
                    company = extract_company_from_slug(href)

            if is_roundup(title) or is_roundup(href):
                continue
            if is_b2b(title) or is_b2b(href):
                continue

            round_type = extract_round(slug + " " + title)
            amount = extract_amount(title)
            full_url = "https://entrackr.com" + href

            if company and is_valid_company_name(company):
                results.append({
                    "company": company,
                    "sector": "Consumer Tech",
                    "round": round_type,
                    "amount": amount,
                    "date": datetime.now().strftime("%Y-%m-%d"),
                    "website": "",
                    "source_url": full_url,
                })

    except Exception as e:
        print(f"  Entrackr error: {e}")

    seen = set()
    unique = [r for r in results if not (r["company"].lower() in seen or seen.add(r["company"].lower()))]
    print(f"  Found {len(unique)} from Entrackr")
    return unique


# ── INC42 ─────────────────────────────────────────────────────────────────────

def scrape_inc42():
    print("Scraping Inc42...")
    results = []
    urls = ["https://inc42.com/buzz/", "https://inc42.com/category/startups/"]

    for url in urls:
        try:
            resp = requests.get(url, headers=HEADERS, timeout=15)
            soup = BeautifulSoup(resp.text, "html.parser")

            for tag in soup.find_all(["h2", "h3"]):
                title = tag.get_text(strip=True)

                if not title or len(title) < 20:
                    continue
                if is_roundup(title):
                    continue
                if not is_funding_article(title):
                    continue
                if is_b2b(title):
                    continue

                a_tag = tag.find("a")
                if not a_tag:
                    parent = tag.parent
                    a_tag = parent.find("a") if parent else None

                href = a_tag["href"] if a_tag and a_tag.get("href") else ""
                full_url = href if href.startswith("http") else "https://inc42.com" + href

                round_type = extract_round(title)
                amount = extract_amount(title)
                company = extract_company_from_title(title)

                if company and is_valid_company_name(company):
                    results.append({
                        "company": company,
                        "sector": "Consumer Tech",
                        "round": round_type if round_type != "Unknown" else "Series A/B",
                        "amount": amount,
                        "date": datetime.now().strftime("%Y-%m-%d"),
                        "website": "",
                        "source_url": full_url,
                    })

            time.sleep(2)

        except Exception as e:
            print(f"  Inc42 error ({url}): {e}")

    seen = set()
    unique = [r for r in results if not (r["company"].lower() in seen or seen.add(r["company"].lower()))]
    print(f"  Found {len(unique)} from Inc42")
    return unique


# ── GOOGLE NEWS RSS ───────────────────────────────────────────────────────────

def scrape_google_news():
    print("Scraping Google News RSS...")
    results = []
    queries = [
        "India startup Series A funding raises",
        "India startup Series B funding raises",
        "Indian startup secures Series A",
        "Indian startup secures Series B",
    ]

    for query in queries:
        try:
            encoded = query.replace(" ", "+")
            rss_url = f"https://news.google.com/rss/search?q={encoded}&hl=en-IN&gl=IN&ceid=IN:en"
            resp = requests.get(rss_url, headers=HEADERS, timeout=15)
            soup = BeautifulSoup(resp.text, "lxml-xml")

            for item in soup.find_all("item")[:25]:
                title_tag = item.find("title")
                link_tag = item.find("link")
                pub_tag = item.find("pubDate")

                if not title_tag:
                    continue

                title_text = title_tag.get_text(strip=True)
                link_text = link_tag.get_text(strip=True) if link_tag else ""
                date_text = pub_tag.get_text(strip=True)[:10] if pub_tag else datetime.now().strftime("%Y-%m-%d")

                # Skip roundup articles
                if is_roundup(title_text):
                    continue
                if not is_funding_article(title_text):
                    continue
                if is_b2b(title_text):
                    continue

                round_type = extract_round(title_text)
                # For Google News, only include if we can confirm Series A or B
                if round_type == "Unknown":
                    continue

                amount = extract_amount(title_text)
                company = extract_company_from_title(title_text)

                if company and is_valid_company_name(company):
                    results.append({
                        "company": company,
                        "sector": "Consumer Tech",
                        "round": round_type,
                        "amount": amount,
                        "date": date_text,
                        "website": "",
                        "source_url": link_text,
                    })

            time.sleep(1)

        except Exception as e:
            print(f"  Google News error: {e}")

    seen = set()
    unique = [r for r in results if not (r["company"].lower() in seen or seen.add(r["company"].lower()))]
    print(f"  Found {len(unique)} from Google News RSS")
    return unique


# ── MAIN ──────────────────────────────────────────────────────────────────────

def run_scraper():
    print("\n=== SCRAPER STARTED ===")
    print(f"Time: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n")

    all_results = []
    all_results += scrape_entrackr()
    all_results += scrape_inc42()
    all_results += scrape_google_news()

    print(f"\nTotal raw results: {len(all_results)}")
    added = save_to_excel(all_results)
    print(f"New rows added to Excel: {added}")
    print("=== SCRAPER DONE ===\n")
    return all_results

if __name__ == "__main__":
    run_scraper()