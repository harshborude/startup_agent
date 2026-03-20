import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

FILE = "startups.xlsx"

def setup_excel():
    if os.path.exists(FILE):
        print(f"{FILE} already exists. Skipping creation.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Startups"

    headers = [
        "Company Name",
        "Sector",
        "Funding Round",
        "Amount Raised",
        "Date Funded",
        "Founder Name",
        "Founder Email",
        "Website",
        "Source URL",
        "Email Status",
        "Email Sent Date",
        "Reply Date",
        "Notes"
    ]

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4B0082", end_color="4B0082", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Column widths
    col_widths = {
        1: 25,   # Company Name
        2: 15,   # Sector
        3: 15,   # Funding Round
        4: 15,   # Amount Raised
        5: 15,   # Date Funded
        6: 20,   # Founder Name
        7: 30,   # Founder Email
        8: 30,   # Website
        9: 40,   # Source URL
        10: 18,  # Email Status
        11: 18,  # Email Sent Date
        12: 15,  # Reply Date
        13: 30,  # Notes
    }

    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 30

    # Freeze the header row
    ws.freeze_panes = "A2"

    wb.save(FILE)
    print(f"Created {FILE} with {len(headers)} columns.")
    print("You can find it at: " + os.path.abspath(FILE))

if __name__ == "__main__":
    setup_excel()